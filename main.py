"""
main.py
-------
Główny skrypt projektu KSeF API 2.0.
Pobiera faktury wystawione i otrzymane, zapisuje do Excel.

Uruchomienie:
    python main.py

Wymagana konfiguracja w pliku .env (skopiuj z .env.example).
"""

import logging
import os
import sys
import time
from pathlib import Path

from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ksef_auth import KSeFAuth, KSeFAuthError
from ksef_invoices import KSeFInvoices, KSeFInvoiceError

# ------------------------------------------------------------------
# Konfiguracja logowania
# ------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ------------------------------------------------------------------
# Wczytaj .env
# ------------------------------------------------------------------
load_dotenv(Path(__file__).parent / ".env")

NIP           = os.getenv("KSEF_NIP", "").strip()
TOKEN         = os.getenv("KSEF_TOKEN", "").strip()
ENV           = os.getenv("KSEF_ENV", "test").strip().lower()
DATE_FROM_STR = os.getenv("DATE_FROM", "2025-01-01").strip()
DATE_TO_STR   = os.getenv("DATE_TO",   "2025-12-31").strip()
PAGE_SIZE     = int(os.getenv("PAGE_SIZE", "100"))

BASE_URLS = {
    "test": "https://api-test.ksef.mf.gov.pl/api/v2",
    "prod": "https://api.ksef.mf.gov.pl/api/v2",
}
BASE_URL = BASE_URLS.get(ENV, BASE_URLS["test"])

OUTPUT_FILE = Path(__file__).parent / "faktury_ksef.xlsx"


# ------------------------------------------------------------------
# Kolumny które chcemy pokazać w Excelu
# Klucze odpowiadają polom zwracanym przez KSeF API
# ------------------------------------------------------------------
INVOICE_COLUMNS = [
    ("_typ",          "Typ faktury"),
    ("ksefNumber",    "Numer KSeF"),
    ("invoiceNumber", "Numer faktury"),
    ("invoiceType",   "Rodzaj faktury"),
    ("issueDate",     "Data wystawienia"),
    ("invoicingDate", "Data przyjęcia w KSeF"),
    ("seller_name",   "Wystawca — nazwa"),
    ("seller_nip",    "Wystawca — NIP"),
    ("buyer_name",    "Nabywca — nazwa"),
    ("buyer_nip",     "Nabywca — NIP"),
    ("netAmount",     "Kwota netto"),
    ("vatAmount",     "Kwota VAT"),
    ("grossAmount",   "Kwota brutto"),
    ("currency",      "Waluta"),
]


def flatten_invoice(inv: dict) -> dict:
    """
    Spłaszcza zagnieżdżone pola faktury do jednopoziomowego słownika.
    KSeF API 2.0 zwraca seller.name, seller.nip, buyer.name, buyer.identifier.value itp.
    """
    flat: dict = {}

    for key, _ in INVOICE_COLUMNS:
        if key == "seller_name":
            flat[key] = (inv.get("seller") or {}).get("name", "")
        elif key == "seller_nip":
            flat[key] = (inv.get("seller") or {}).get("nip", "")
        elif key == "buyer_name":
            flat[key] = (inv.get("buyer") or {}).get("name", "")
        elif key == "buyer_nip":
            # buyer.identifier.value (NIP nabywcy zagnieżdżony głębiej)
            flat[key] = ((inv.get("buyer") or {}).get("identifier") or {}).get("value", "")
        else:
            flat[key] = inv.get(key, "")

    return flat


# ------------------------------------------------------------------
# Zapis do Excel
# ------------------------------------------------------------------
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")  # ciemnoniebieski
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
ALT_ROW_FILL = PatternFill("solid", fgColor="D6E4F0")  # jasnobłękitny


def write_sheet(ws, invoices: list[dict], title: str) -> None:
    """Zapisuje listę faktur do arkusza Excel z formatowaniem."""

    col_keys   = [c[0] for c in INVOICE_COLUMNS]
    col_labels = [c[1] for c in INVOICE_COLUMNS]

    # Nagłówek
    ws.append(col_labels)
    for col_idx, _ in enumerate(col_labels, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 28

    # Dane
    for row_idx, inv in enumerate(invoices, start=2):
        flat = flatten_invoice(inv)
        for col_idx, key in enumerate(col_keys, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = flat.get(key, "")
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

    # Szerokości kolumn
    column_widths = [28, 36, 22, 18, 18, 22, 36, 16, 36, 16, 14, 14, 14, 10]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Zamrożenie nagłówka
    ws.freeze_panes = "A2"

    # Autofiltr
    ws.auto_filter.ref = ws.dimensions


def save_to_excel(wystawione: list[dict], otrzymane: list[dict], path: Path) -> None:
    """Tworzy plik Excel z dwoma arkuszami."""
    wb = openpyxl.Workbook()

    # Arkusz 1 — Wystawione
    ws1 = wb.active
    ws1.title = "Wystawione"
    write_sheet(ws1, wystawione, "Wystawione")

    # Arkusz 2 — Otrzymane
    ws2 = wb.create_sheet("Otrzymane")
    write_sheet(ws2, otrzymane, "Otrzymane")

    # Arkusz 3 — Podsumowanie
    ws3 = wb.create_sheet("Podsumowanie")
    ws3.append(["KSeF API 2.0 — Pobieranie faktur"])
    ws3.append([])
    ws3.append(["NIP firmy:",       NIP])
    ws3.append(["Środowisko:",      ENV.upper()])
    ws3.append(["Zakres dat:",      f"{DATE_FROM_STR} — {DATE_TO_STR}"])
    ws3.append(["Faktur wystawionych:", len(wystawione)])
    ws3.append(["Faktur otrzymanych:",  len(otrzymane)])
    ws3.append(["Łącznie:",             len(wystawione) + len(otrzymane)])

    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 30
    ws3["A1"].font = Font(bold=True, size=14)

    wb.save(path)


# ------------------------------------------------------------------
# MAIN
# ------------------------------------------------------------------
def main() -> None:
    logger.info("=" * 55)
    logger.info("  KSeF API 2.0 — Pobieranie faktur")
    logger.info("=" * 55)

    # Walidacja konfiguracji
    if not NIP or not TOKEN:
        logger.error(
            "Brak konfiguracji! Uzupełnij plik .env (KSEF_NIP i KSEF_TOKEN).\n"
            "Skopiuj .env.example → .env i uzupełnij wartości."
        )
        sys.exit(1)

    logger.info(f"NIP: {NIP} | Środowisko: {ENV.upper()} | Zakres: {DATE_FROM_STR} — {DATE_TO_STR}")

    # ------------------------------------------------------------------
    # 1. Uwierzytelnienie
    # ------------------------------------------------------------------
    auth = KSeFAuth(nip=NIP, ksef_token=TOKEN, env=ENV)
    try:
        auth.authenticate()
    except KSeFAuthError as e:
        logger.error(f"Błąd uwierzytelnienia: {e}")
        sys.exit(1)

    auth_headers = auth.get_auth_headers()
    time.sleep(1)  # krótkie opóźnienie po uwierzytelnieniu

    # ------------------------------------------------------------------
    # 2. Przygotowanie dat w formacie ISO 8601
    # ------------------------------------------------------------------
    date_from = KSeFInvoices.to_iso(DATE_FROM_STR, end_of_day=False)
    date_to   = KSeFInvoices.to_iso(DATE_TO_STR,   end_of_day=True)
    logger.info(f"Zakres dat ISO: {date_from}  →  {date_to}")

    # ------------------------------------------------------------------
    # 3. Pobieranie faktur
    # ------------------------------------------------------------------
    # auth przekazany do klienta — obsłuży wygaśnięcie tokena (401) automatycznie
    client = KSeFInvoices(base_url=BASE_URL, auth_headers=auth_headers, page_size=PAGE_SIZE, auth=auth)

    wystawione: list[dict] = []
    otrzymane:  list[dict] = []

    try:
        logger.info("\n--- FAKTURY WYSTAWIONE ---")
        wystawione = client.fetch_all("Subject1", date_from, date_to)

        logger.info("\n--- FAKTURY OTRZYMANE ---")
        otrzymane = client.fetch_all("Subject2", date_from, date_to)

    except KSeFInvoiceError as e:
        logger.error(f"Błąd pobierania faktur: {e}")
        sys.exit(1)

    # ------------------------------------------------------------------
    # 4. Zapis do Excel
    # ------------------------------------------------------------------
    logger.info(f"\nZapisywanie do pliku: {OUTPUT_FILE.name} ...")
    save_to_excel(wystawione, otrzymane, OUTPUT_FILE)

    logger.info("\n" + "=" * 55)
    logger.info(f"  ✓ Gotowe! Plik zapisany: {OUTPUT_FILE}")
    logger.info(f"  Faktur wystawionych: {len(wystawione)}")
    logger.info(f"  Faktur otrzymanych:  {len(otrzymane)}")
    logger.info(f"  Łącznie:             {len(wystawione) + len(otrzymane)}")
    logger.info("=" * 55)


if __name__ == "__main__":
    main()
